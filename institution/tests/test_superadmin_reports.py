"""
Tests for Story 2.5 — Cross-Institution Aggregate Reports.

Covers:
- SUPERADMIN-only access (GET and POST)
- per_institution_aggregate() — scoped workbook (AC: #1)
- cross_institution_aggregate() — Network Summary + per-institution sheets (AC: #2)
- BasePDFGenerator institution parameter (AC: #3)
- Zero-patient institution handling (AC: #4)
"""
import os

from django.test import TestCase, Client, override_settings
from django.urls import reverse
from django.contrib.auth import get_user_model

from institution.models import Institution
from ndas.custom_codes.choice import UserType, SubscriptionStatus

User = get_user_model()


def make_superadmin(username='superadmin'):
    return User.objects.create_user(
        username=username, password='testpass123',
        email=f'{username}@test.com', first_name='Super',
        position='Medical Officer', mobile_primary='0771234567',
        user_type=UserType.SUPERADMIN, is_superuser=True, is_staff=True,
    )


def make_admin(username='admin', institution=None):
    return User.objects.create_user(
        username=username, password='testpass123',
        email=f'{username}@test.com', first_name='Admin',
        position='Medical Officer', mobile_primary='0771234568',
        user_type=UserType.ADMIN, institution=institution,
    )


STATIC_OVERRIDE = override_settings(
    STORAGES={"staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"}},
    RATELIMIT_ENABLE=False,
    MULTI_INSTITUTION_ENABLED=True,
)


@STATIC_OVERRIDE
class SuperadminReportsAccessTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.url = reverse('institution:superadmin-reports')
        self.superadmin = make_superadmin()
        self.institution = Institution.objects.create(name='Test Hosp', slug='test-hosp')

    def test_unauthenticated_redirects(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('login', response['Location'].lower())

    def test_superadmin_get_renders_form(self):
        self.client.login(username='superadmin', password='testpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'institution/superadmin_reports.html')
        self.assertIn('institutions', response.context)

    def test_non_superadmin_redirected(self):
        """ADMIN must be redirected — 200 would mean a security bypass."""
        make_admin(institution=self.institution)
        self.client.login(username='admin', password='testpass123')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302,
            "ADMIN must be redirected from superadmin reports (not 200)")


@STATIC_OVERRIDE
class SuperadminReportsValidationTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.url = reverse('institution:superadmin-reports')
        self.superadmin = make_superadmin()
        self.institution = Institution.objects.create(name='Export Hospital', slug='export-hosp')
        self.client.login(username='superadmin', password='testpass123')

    def test_invalid_scope_shows_error(self):
        response = self.client.post(self.url, {'scope': 'invalid_scope'})
        self.assertEqual(response.status_code, 200)
        msgs = list(response.context['messages'])
        self.assertTrue(any('Invalid report scope' in str(m) for m in msgs))

    def test_per_institution_without_institution_id_shows_error(self):
        response = self.client.post(self.url, {'scope': 'per_institution', 'institution_id': ''})
        self.assertEqual(response.status_code, 200)
        msgs = list(response.context['messages'])
        self.assertTrue(any('select' in str(m).lower() for m in msgs))


@STATIC_OVERRIDE
class PerInstitutionExcelTest(TestCase):
    """AC #1: per_institution_aggregate() generates an institution-scoped workbook."""

    def setUp(self):
        self.client = Client()
        self.url = reverse('institution:superadmin-reports')
        self.superadmin = make_superadmin()
        self.institution = Institution.objects.create(
            name='Alpha Hospital', slug='alpha-hospital',
            subscription_status=SubscriptionStatus.ACTIVE,
        )
        self.client.login(username='superadmin', password='testpass123')

    def test_per_institution_export_returns_xlsx(self):
        """AC #1: POST with scope=per_institution returns Excel file attachment."""
        response = self.client.post(self.url, {
            'scope': 'per_institution',
            'institution_id': str(self.institution.pk),
        })
        self.assertEqual(response.status_code, 200,
            "per_institution export must return 200 with file content")
        self.assertIn('spreadsheetml.sheet', response.get('Content-Type', ''),
            "Content-Type must be Excel")
        self.assertIn('attachment', response.get('Content-Disposition', ''),
            "Must be a file download attachment")

    def test_per_institution_filename_contains_slug(self):
        """AC #1: Downloaded filename includes the institution slug."""
        response = self.client.post(self.url, {
            'scope': 'per_institution',
            'institution_id': str(self.institution.pk),
        })
        self.assertIn('alpha-hospital', response.get('Content-Disposition', ''),
            "Filename must contain the institution slug")

    def test_per_institution_aggregate_method_exists(self):
        """AC #1: ExcelReportGenerator has per_institution_aggregate() method."""
        from reports.utils.excel_generator import ExcelReportGenerator
        self.assertTrue(hasattr(ExcelReportGenerator(), 'per_institution_aggregate'),
            "ExcelReportGenerator must have per_institution_aggregate() method")

    def test_per_institution_workbook_has_patients_sheet(self):
        """AC #1: per_institution_aggregate() workbook contains a Patients sheet."""
        from reports.utils.excel_generator import ExcelReportGenerator
        from openpyxl import load_workbook
        gen = ExcelReportGenerator()
        output_path, _ = gen.per_institution_aggregate(institution=self.institution)
        try:
            wb = load_workbook(output_path)
            self.assertIn('Patients', wb.sheetnames,
                "AC #1: Workbook must contain a 'Patients' sheet")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


@STATIC_OVERRIDE
class CrossInstitutionExcelTest(TestCase):
    """AC #2: cross_institution_aggregate() — Network Summary + per-institution sheets."""

    def setUp(self):
        self.client = Client()
        self.url = reverse('institution:superadmin-reports')
        self.superadmin = make_superadmin()
        self.inst_a = Institution.objects.create(name='Alpha Hospital', slug='alpha')
        self.inst_b = Institution.objects.create(name='Beta Clinic', slug='beta')
        self.client.login(username='superadmin', password='testpass123')

    def test_cross_institution_export_returns_xlsx(self):
        """AC #2: POST with scope=cross_institution returns Excel file."""
        response = self.client.post(self.url, {'scope': 'cross_institution'})
        self.assertEqual(response.status_code, 200,
            "cross_institution export must return 200 with file content")
        self.assertIn('spreadsheetml.sheet', response.get('Content-Type', ''),
            "Content-Type must be Excel")

    def test_cross_institution_aggregate_method_exists(self):
        """AC #2: ExcelReportGenerator has cross_institution_aggregate() method."""
        self.assertTrue(hasattr(__import__('reports.utils.excel_generator',
            fromlist=['ExcelReportGenerator']).ExcelReportGenerator(),
            'cross_institution_aggregate'),
            "ExcelReportGenerator must have cross_institution_aggregate() method")

    def test_cross_institution_workbook_has_network_summary_sheet(self):
        """AC #2: Workbook contains 'Network Summary' sheet."""
        from reports.utils.excel_generator import ExcelReportGenerator
        from openpyxl import load_workbook
        gen = ExcelReportGenerator()
        output_path, _ = gen.cross_institution_aggregate()
        try:
            wb = load_workbook(output_path)
            self.assertIn('Network Summary', wb.sheetnames,
                "AC #2: Workbook must contain 'Network Summary' sheet")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_cross_institution_workbook_has_per_institution_sheets(self):
        """AC #2: Workbook contains one breakdown sheet per institution (Sheets 2..N)."""
        from reports.utils.excel_generator import ExcelReportGenerator
        from openpyxl import load_workbook
        gen = ExcelReportGenerator()
        output_path, _ = gen.cross_institution_aggregate()
        try:
            wb = load_workbook(output_path)
            # Network Summary + at least 2 institution sheets
            self.assertGreaterEqual(len(wb.sheetnames), 3,
                "AC #2: Workbook must have Network Summary + one sheet per institution "
                f"(got {wb.sheetnames})")
            # Institution name (truncated to 31 chars) must appear as sheet name
            sheet_names = wb.sheetnames
            self.assertTrue(
                any('Alpha' in s for s in sheet_names),
                f"AC #2: Alpha Hospital breakdown sheet must exist (sheets: {sheet_names})"
            )
            self.assertTrue(
                any('Beta' in s for s in sheet_names),
                f"AC #2: Beta Clinic breakdown sheet must exist (sheets: {sheet_names})"
            )
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_network_summary_contains_all_institutions(self):
        """AC #2: Network Summary sheet has a row for each institution."""
        from reports.utils.excel_generator import ExcelReportGenerator
        from openpyxl import load_workbook
        gen = ExcelReportGenerator()
        output_path, _ = gen.cross_institution_aggregate()
        try:
            wb = load_workbook(output_path)
            ws = wb['Network Summary']
            cell_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
            self.assertTrue(any('Alpha' in str(v) for v in cell_values if v),
                "Network Summary must contain Alpha Hospital row")
            self.assertTrue(any('Beta' in str(v) for v in cell_values if v),
                "Network Summary must contain Beta Clinic row")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


@STATIC_OVERRIDE
class PDFInstitutionBrandingTest(TestCase):
    """AC #3: BasePDFGenerator accepts institution parameter for branding."""

    def setUp(self):
        self.superadmin = make_superadmin()
        self.institution = Institution.objects.create(
            name='Branded Hospital', slug='branded-hosp',
            subscription_status=SubscriptionStatus.ACTIVE,
        )

    def test_base_pdf_generator_accepts_institution_param(self):
        """AC #3: BasePDFGenerator stores institution as self.institution."""
        from reports.utils.pdf_generator import BasePDFGenerator
        try:
            gen = BasePDFGenerator(institution=self.institution)
            self.assertEqual(gen.institution, self.institution,
                "BasePDFGenerator must store institution as self.institution")
        except TypeError as e:
            self.fail(f"AC #3: BasePDFGenerator must accept institution= parameter: {e}")

    def test_base_pdf_generator_default_institution_is_none(self):
        """AC #3: institution=None preserves backward compatibility."""
        from reports.utils.pdf_generator import BasePDFGenerator
        gen = BasePDFGenerator()
        self.assertIsNone(gen.institution,
            "Default institution must be None (backward compatible)")

    def test_patient_pdf_generator_inherits_institution_param(self):
        """AC #3: PatientPDFGenerator accepts institution= (inherits from BasePDFGenerator)."""
        from reports.utils.pdf_generator import PatientPDFGenerator
        try:
            gen = PatientPDFGenerator(institution=self.institution)
            self.assertEqual(gen.institution, self.institution)
        except TypeError as e:
            self.fail(f"PatientPDFGenerator must accept institution= parameter: {e}")


@STATIC_OVERRIDE
class ZeroPatientInstitutionTest(TestCase):
    """AC #4: Zero-patient institutions handled without exceptions."""

    def setUp(self):
        self.superadmin = make_superadmin()
        self.inst_empty = Institution.objects.create(
            name='Empty Clinic', slug='empty-clinic',
            subscription_status=SubscriptionStatus.GRACE,
        )
        self.inst_also_empty = Institution.objects.create(
            name='Also Empty', slug='also-empty',
            subscription_status=SubscriptionStatus.EXPIRED,
        )

    def test_per_institution_zero_patients_no_exception(self):
        """AC #4: per_institution_aggregate() with zero patients does not raise."""
        from reports.utils.excel_generator import ExcelReportGenerator
        gen = ExcelReportGenerator()
        output_path = None
        try:
            output_path, metadata = gen.per_institution_aggregate(institution=self.inst_empty)
            self.assertIsNotNone(output_path,
                "AC #4: Must return output path even with zero patients")
        except Exception as e:
            self.fail(f"AC #4: per_institution_aggregate must not raise for zero patients: {e}")
        finally:
            if output_path and os.path.exists(output_path):
                os.unlink(output_path)

    def test_cross_institution_zero_patients_no_exception(self):
        """AC #4: cross_institution_aggregate() with zero-patient institutions does not raise."""
        from reports.utils.excel_generator import ExcelReportGenerator
        gen = ExcelReportGenerator()
        output_path = None
        try:
            output_path, metadata = gen.cross_institution_aggregate()
        except Exception as e:
            self.fail(f"AC #4: cross_institution_aggregate must not raise with zero patients: {e}")
        finally:
            if output_path and os.path.exists(output_path):
                os.unlink(output_path)

    def test_cross_institution_zero_patients_shown_as_zeros(self):
        """AC #4: Zero-patient institutions appear in Network Summary with 0 counts."""
        from reports.utils.excel_generator import ExcelReportGenerator
        from openpyxl import load_workbook
        gen = ExcelReportGenerator()
        output_path, _ = gen.cross_institution_aggregate()
        try:
            wb = load_workbook(output_path, data_only=True)
            ws = wb['Network Summary']
            found_empty = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and 'Empty Clinic' in str(row[0]):
                    found_empty = True
                    patient_col = row[4]  # Column 5 = Total Patients
                    self.assertEqual(patient_col, 0,
                        "AC #4: Zero-patient institution must show 0 in patient count column")
                    break
            self.assertTrue(found_empty,
                "AC #4: Empty Clinic must appear in Network Summary sheet")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
