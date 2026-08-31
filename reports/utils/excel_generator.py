"""
Excel Report Generation Utilities

This module contains Excel generator classes for creating data export reports
using openpyxl for research and statistical analysis.
"""

import logging
import os
import uuid
from datetime import datetime

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ndas.custom_codes.custom_methods import escape_excel_row
from patients.models import (
    Patient, GMAssessment, HINEAssessment,
    DevelopmentalAssessment, CDICRecord, GeneralPaediatricAssessment
)

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel reports for research data exports"""

    def __init__(self):
        """Initialize Excel generator"""
        self.workbook = None

    @staticmethod
    def make_naive(dt):
        """
        Convert timezone-aware datetime to naive datetime for Excel compatibility

        Args:
            dt: datetime object (can be aware or naive)

        Returns:
            Naive datetime object or original value if not datetime
        """
        if dt is None:
            return None
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            # Convert to local time and remove timezone info
            return timezone.localtime(dt).replace(tzinfo=None)
        return dt

    @staticmethod
    def sanitize_row(row):
        """
        Escape spreadsheet formula injection in every cell of a data row
        before ws.append(). See ndas.custom_codes.custom_methods.escape_excel_row.
        """
        return escape_excel_row(row)

    @staticmethod
    def anonymize_value(value, field_name, patient_id):
        """
        Anonymize patient identifier fields for research compliance

        Args:
            value: Original field value
            field_name: Name of the field being anonymized
            patient_id: Patient ID for generating consistent anonymous IDs

        Returns:
            Anonymized value or original value if field doesn't need anonymization
        """
        # Fields that should be anonymized
        anonymizable_fields = ['baby_name', 'mother_name', 'bht', 'nnc_no', 'ptc_no', 'pc_no', 'pin', 'disk_no']

        if field_name in anonymizable_fields:
            if field_name == 'baby_name':
                return f'PATIENT_{patient_id:03d}'
            elif field_name == 'mother_name':
                return f'MOTHER_{patient_id:03d}'
            elif field_name in ['bht', 'nnc_no', 'ptc_no', 'pc_no', 'pin', 'disk_no']:
                return f'ANON_{field_name.upper()}_{patient_id:03d}'

        return value

    def get_field_header_map(self):
        """
        Get mapping of field names to human-readable headers

        Returns:
            dict: Mapping of field names to display headers
        """
        return {
            'id': 'ID',
            'baby_name': 'Baby Name',
            'mother_name': 'Mother Name',
            'bht': 'BHT',
            'nnc_no': 'NNC',
            'ptc_no': 'PTC',
            'pc_no': 'PC',
            'pin': 'PIN',
            'disk_no': 'Disk No',
            'sex': 'Gender',
            'dob_tob': 'Date/Time of Birth',
            'pog_wks': 'Gestational Age (weeks)',
            'pog_days': 'Gestational Age (days)',
            'mode_of_delivery': 'Mode of Delivery',
            'apgar_1': 'APGAR 1min',
            'apgar_5': 'APGAR 5min',
            'apgar_10': 'APGAR 10min',
            'resuscitated': 'Resuscitated',
            'created_at': 'Created At',
            'updated_at': 'Updated At',
        }

    def apply_advanced_filters(self, queryset, filters, model_type='patient'):
        """
        Apply advanced filters to queryset

        Args:
            queryset: Django queryset to filter
            filters: Dict of filter parameters
            model_type: Type of model ('patient', 'gm', 'hine', etc.)

        Returns:
            Filtered queryset
        """
        if not filters:
            return queryset

        # Apply patient-level filters
        if model_type == 'patient':
            # Gestational age filter
            if 'pog_min' in filters:
                queryset = queryset.filter(pog_wks__gte=filters['pog_min'])
            if 'pog_max' in filters:
                queryset = queryset.filter(pog_wks__lte=filters['pog_max'])

            # APGAR score filter
            if 'apgar_timepoint' in filters:
                timepoint = filters['apgar_timepoint']
                field_name = f'apgar_{timepoint}'
                if 'apgar_min' in filters:
                    queryset = queryset.filter(**{f'{field_name}__gte': filters['apgar_min']})
                if 'apgar_max' in filters:
                    queryset = queryset.filter(**{f'{field_name}__lte': filters['apgar_max']})

            # Gender filter
            if 'gender' in filters:
                queryset = queryset.filter(gender__in=filters['gender'])

            # Resuscitation status filter
            if 'resuscitation_status' in filters:
                resus_values = []
                if 'yes' in filters['resuscitation_status']:
                    resus_values.append(True)
                if 'no' in filters['resuscitation_status']:
                    resus_values.append(False)
                if resus_values:
                    queryset = queryset.filter(resuscitated__in=resus_values)

        # Apply GM assessment filters
        elif model_type == 'gm':
            if 'gm_diagnosis' in filters:
                queryset = queryset.filter(diagnosis_conclusion__in=filters['gm_diagnosis'])

            # Also filter by patient criteria if specified
            patient_filters = {}
            if 'pog_min' in filters:
                patient_filters['patient__pog_wks__gte'] = filters['pog_min']
            if 'pog_max' in filters:
                patient_filters['patient__pog_wks__lte'] = filters['pog_max']
            if 'gender' in filters:
                patient_filters['patient__gender__in'] = filters['gender']
            if patient_filters:
                queryset = queryset.filter(**patient_filters)

        # Apply HINE assessment filters
        elif model_type == 'hine':
            if 'hine_score_min' in filters:
                queryset = queryset.filter(score__gte=filters['hine_score_min'])
            if 'hine_score_max' in filters:
                queryset = queryset.filter(score__lte=filters['hine_score_max'])

            # Also filter by patient criteria if specified
            patient_filters = {}
            if 'pog_min' in filters:
                patient_filters['patient__pog_wks__gte'] = filters['pog_min']
            if 'pog_max' in filters:
                patient_filters['patient__pog_wks__lte'] = filters['pog_max']
            if 'gender' in filters:
                patient_filters['patient__gender__in'] = filters['gender']
            if patient_filters:
                queryset = queryset.filter(**patient_filters)

        # Apply filters to other assessment types (filter by patient criteria)
        elif model_type in ['developmental', 'cdic', 'gpa']:
            patient_filters = {}
            if 'pog_min' in filters:
                patient_filters['patient__pog_wks__gte'] = filters['pog_min']
            if 'pog_max' in filters:
                patient_filters['patient__pog_wks__lte'] = filters['pog_max']
            if 'gender' in filters:
                patient_filters['patient__gender__in'] = filters['gender']
            if patient_filters:
                queryset = queryset.filter(**patient_filters)

        return queryset

    def create_workbook(self):
        """Initialize new workbook"""
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def calculate_quality_metrics(self, start_date=None, end_date=None, institution=None):
        """
        Calculate data quality metrics for the export

        Returns:
            dict: Quality metrics with status and notes
        """
        from django.db.models import Count, Q
        from datetime import datetime, time

        metrics = {}

        # Scope all queries to institution (None = all institutions, Phase 1 safe)
        _pts = Patient.objects.for_institution(institution)

        # Convert date objects to timezone-aware datetime for filtering
        start_datetime = None
        end_datetime = None
        if start_date:
            # Make timezone-aware datetime at start of day
            start_datetime = timezone.make_aware(datetime.combine(start_date, time.min))
        if end_date:
            # Make timezone-aware datetime at end of day
            end_datetime = timezone.make_aware(datetime.combine(end_date, time.max))

        # Calculate patient data completeness
        patient_filter = Q()
        if start_datetime:
            patient_filter &= Q(created_at__gte=start_datetime)
        if end_datetime:
            patient_filter &= Q(created_at__lte=end_datetime)

        total_patients = _pts.filter(patient_filter).count()
        if total_patients > 0:
            # Check for patients with complete birth data (using actual fields)
            complete_birth_data = _pts.filter(
                patient_filter,
                pog_wks__isnull=False,
                apgar_1__isnull=False,
                apgar_5__isnull=False,
                dob_tob__isnull=False
            ).count()

            completeness_pct = (complete_birth_data / total_patients * 100)
            metrics['Patient Birth Data Completeness'] = {
                'value': f'{completeness_pct:.1f}%',
                'status': 'Good' if completeness_pct >= 80 else 'Warning' if completeness_pct >= 50 else 'Critical',
                'notes': f'{complete_birth_data}/{total_patients} patients have complete birth data'
            }

            # Check for patients with assessments
            patients_with_assessments = _pts.filter(
                patient_filter
            ).filter(
                Q(gm_assessments__isnull=False) |
                Q(hine_assessments__isnull=False) |
                Q(developmental_assessments__isnull=False) |
                Q(cdic_records__isnull=False) |
                Q(gpa_assessments__isnull=False)
            ).distinct().count()

            assessment_coverage = (patients_with_assessments / total_patients * 100) if total_patients > 0 else 0
            metrics['Patient Assessment Coverage'] = {
                'value': f'{assessment_coverage:.1f}%',
                'status': 'Good' if assessment_coverage >= 60 else 'Warning' if assessment_coverage >= 30 else 'Critical',
                'notes': f'{patients_with_assessments}/{total_patients} patients have at least one assessment'
            }

        # Assessment data quality
        gm_total = GMAssessment.objects.filter(patient__in=_pts).count()
        if gm_total > 0:
            gm_complete = GMAssessment.objects.filter(
                patient__in=_pts,
                diagnosis_conclusion__isnull=False,
                date_of_assessment__isnull=False
            ).count()
            gm_quality = (gm_complete / gm_total * 100)
            metrics['GM Assessment Data Quality'] = {
                'value': f'{gm_quality:.1f}%',
                'status': 'Good' if gm_quality >= 90 else 'Warning',
                'notes': f'{gm_complete}/{gm_total} GM assessments have complete data'
            }

        return metrics

    def style_header_row(self, worksheet, row_num=1):
        """Apply styling to header row"""
        header_fill = PatternFill(
            start_color='4472C4',
            end_color='4472C4',
            fill_type='solid'
        )
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def auto_adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content (optimized)"""
        # Optimized: Calculate max width from header + sample rows instead of all rows
        for idx, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(idx)

            # Check header (row 1)
            header_cell = worksheet.cell(row=1, column=idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))

            # Sample first 100 rows for width calculation (performance optimization)
            sample_size = min(100, worksheet.max_row)
            for row_num in range(2, sample_size + 1):
                cell = worksheet.cell(row=row_num, column=idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def add_patients_sheet(self, workbook, queryset, selected_fields=None, anonymize=False):
        """
        Add comprehensive Patients worksheet with customizable fields and anonymization

        Args:
            workbook: Openpyxl workbook object
            queryset: Patient queryset
            selected_fields: List of field names to include (None = all fields)
            anonymize: Boolean, whether to anonymize patient identifiers
        """
        ws = workbook.create_sheet("Patients")

        # Get field header map
        field_header_map = self.get_field_header_map()

        # Default to all fields if none selected
        if selected_fields is None or len(selected_fields) == 0:
            selected_fields = list(field_header_map.keys())

        # Build headers based on selected fields
        headers = [field_header_map.get(field, field) for field in selected_fields]
        ws.append(headers)

        # Style header
        self.style_header_row(ws)

        # Data rows with selected fields
        for patient in queryset:
            row = []
            for field in selected_fields:
                value = None

                # Get field value
                if field == 'id':
                    value = patient.id
                elif field == 'baby_name':
                    value = patient.baby_name or ''
                elif field == 'mother_name':
                    value = patient.mother_name or ''
                elif field == 'bht':
                    value = patient.bht or ''
                elif field == 'nnc_no':
                    value = patient.nnc_no or ''
                elif field == 'ptc_no':
                    value = patient.ptc_no or ''
                elif field == 'pc_no':
                    value = patient.pc_no or ''
                elif field == 'pin':
                    value = patient.pin or ''
                elif field == 'disk_no':
                    value = patient.disk_no or ''
                elif field == 'sex':
                    value = patient.gender or ''
                elif field == 'dob_tob':
                    value = self.make_naive(patient.dob_tob)
                elif field == 'pog_wks':
                    value = patient.pog_wks
                elif field == 'pog_days':
                    value = patient.pog_days
                elif field == 'mode_of_delivery':
                    value = patient.mo_delivery or ''
                elif field == 'apgar_1':
                    value = patient.apgar_1
                elif field == 'apgar_5':
                    value = patient.apgar_5
                elif field == 'apgar_10':
                    value = patient.apgar_10
                elif field == 'resuscitated':
                    value = patient.resuscitated
                elif field == 'created_at':
                    value = self.make_naive(patient.created_at)
                elif field == 'updated_at':
                    value = self.make_naive(patient.updated_at)
                else:
                    # Try to get attribute dynamically
                    value = getattr(patient, field, '')

                # Apply anonymization if enabled
                if anonymize:
                    value = self.anonymize_value(value, field, patient.id)

                row.append(value)

            ws.append(self.sanitize_row(row))

        # Auto-adjust widths
        self.auto_adjust_column_widths(ws)

    def add_gm_assessments_sheet(self, workbook, queryset):
        """Add GM Assessments worksheet"""
        ws = workbook.create_sheet("GM Assessments")

        # Headers
        headers = [
            'ID', 'Patient BHT', 'Patient Name', 'Assessment Date',
            'Age at Assessment (weeks)', 'Age at Assessment (days)',
            'Classification', 'Created At'
        ]
        ws.append(headers)

        # Style header
        self.style_header_row(ws)

        # Data rows
        for assessment in queryset.select_related('patient'):
            row = [
                assessment.id,
                assessment.patient.bht or '',
                assessment.patient.baby_name or '',
                self.make_naive(assessment.date_of_assessment),
                getattr(assessment, 'age_at_assessment_weeks', ''),
                getattr(assessment, 'age_at_assessment_days', ''),
                getattr(assessment, 'diagnosis_conclusion', ''),
                self.make_naive(assessment.created_at),
            ]
            ws.append(self.sanitize_row(row))

        # Auto-adjust widths
        self.auto_adjust_column_widths(ws)

    def add_hine_assessments_sheet(self, workbook, queryset):
        """Add HINE Assessments worksheet"""
        ws = workbook.create_sheet("HINE Assessments")

        # Headers
        headers = [
            'ID', 'Patient BHT', 'Patient Name', 'Assessment Date',
            'Age at Assessment (weeks)', 'Age at Assessment (days)',
            'Total Score', 'Created At'
        ]
        ws.append(headers)

        # Style header
        self.style_header_row(ws)

        # Data rows
        for assessment in queryset.select_related('patient'):
            row = [
                assessment.id,
                assessment.patient.bht or '',
                assessment.patient.baby_name or '',
                self.make_naive(assessment.date_of_assessment),
                getattr(assessment, 'age_at_assessment_weeks', ''),
                getattr(assessment, 'age_at_assessment_days', ''),
                assessment.score,
                self.make_naive(assessment.created_at),
            ]
            ws.append(self.sanitize_row(row))

        # Auto-adjust widths
        self.auto_adjust_column_widths(ws)

    def add_developmental_assessments_sheet(self, workbook, queryset):
        """Add Developmental Assessments worksheet"""
        ws = workbook.create_sheet("Developmental Assessments")

        # Headers
        headers = [
            'ID', 'Patient BHT', 'Patient Name', 'Assessment Date',
            'GM Age From (months)', 'GM Age To (months)', 'GM Details',
            'FMV Age From (months)', 'FMV Age To (months)', 'FMV Details',
            'Language Age From (months)', 'Language Age To (months)', 'Language Details',
            'Social Age From (months)', 'Social Age To (months)', 'Social Details',
            'Created At', 'Added By'
        ]
        ws.append(headers)
        self.style_header_row(ws)

        # Data rows
        for assessment in queryset.select_related('patient', 'added_by'):
            row = [
                assessment.id,
                assessment.patient.bht or '',
                assessment.patient.baby_name or '',
                self.make_naive(assessment.date_of_assessment),
                assessment.gm_age_from,
                assessment.gm_age_to,
                assessment.gm_details or '',
                assessment.fmv_age_from,
                assessment.fmv_age_to,
                assessment.fmv_details or '',
                assessment.language_age_from if hasattr(assessment, 'language_age_from') else '',
                assessment.language_age_to if hasattr(assessment, 'language_age_to') else '',
                assessment.language_details if hasattr(assessment, 'language_details') else '',
                assessment.social_age_from if hasattr(assessment, 'social_age_from') else '',
                assessment.social_age_to if hasattr(assessment, 'social_age_to') else '',
                assessment.social_details if hasattr(assessment, 'social_details') else '',
                self.make_naive(assessment.created_at),
                assessment.added_by.username if assessment.added_by else '',
            ]
            ws.append(self.sanitize_row(row))

        self.auto_adjust_column_widths(ws)

    def add_cdic_records_sheet(self, workbook, queryset):
        """Add CDIC Records worksheet"""
        ws = workbook.create_sheet("CDIC Records")

        # Headers
        headers = [
            'ID', 'Patient BHT', 'Patient Name', 'Assessment Date',
            'Assessment Done By', 'Assessment', 'Today Interventions',
            'Next Appointment Date', 'Next Appointment Plan',
            'Is Discharged', 'Discharge Date', 'Discharged By',
            'Created At', 'Added By'
        ]
        ws.append(headers)
        self.style_header_row(ws)

        # Data rows
        for record in queryset.select_related('patient', 'added_by'):
            row = [
                record.id,
                record.patient.bht or '',
                record.patient.baby_name or '',
                self.make_naive(record.assessment_date) if record.assessment_date else '',
                record.assessment_done_by or '',
                record.assessment or '',
                record.today_interventions or '',
                self.make_naive(record.next_appointment_date),
                record.next_appointment_plan or '',
                record.is_discharged,
                self.make_naive(record.discharge_date) if record.discharge_date else '',
                record.discharged_by or '',
                self.make_naive(record.created_at),
                record.added_by.username if record.added_by else '',
            ]
            ws.append(self.sanitize_row(row))

        self.auto_adjust_column_widths(ws)

    def add_gpa_assessments_sheet(self, workbook, queryset):
        """Add General Paediatric Assessments worksheet"""
        ws = workbook.create_sheet("GPA Assessments")

        # Headers
        headers = [
            'ID', 'Patient BHT', 'Patient Name', 'Assessment Date',
            'Healthcare Provider', 'Current Problems', 'Physical Examination',
            'Investigation Summary', 'Prescribed Medications', 'Next Plan',
            'Next Assessment Date', 'Created At', 'Added By'
        ]
        ws.append(headers)
        self.style_header_row(ws)

        # Data rows
        for assessment in queryset.select_related('patient', 'added_by'):
            row = [
                assessment.id,
                assessment.patient.bht or '',
                assessment.patient.baby_name or '',
                self.make_naive(assessment.assessment_date),
                assessment.healthcare_provider or '',
                assessment.current_problems or '',
                assessment.physical_examination or '',
                assessment.investigation_summary or '',
                assessment.prescribed_medications or '',
                assessment.next_plan or '',
                self.make_naive(assessment.next_assessment_date),
                self.make_naive(assessment.created_at),
                assessment.added_by.username if assessment.added_by else '',
            ]
            ws.append(self.sanitize_row(row))

        self.auto_adjust_column_widths(ws)

    def add_summary_statistics_sheet(self, workbook, metadata):
        """
        Add statistical summary sheet for researchers

        Args:
            workbook: Openpyxl workbook object
            metadata: Dict containing sheet counts and statistics

        Unlike the other add_*_sheet methods, this one writes cells directly
        (ws['A1'] = ...) instead of via ws.append(self.sanitize_row(row)).
        Safe today because `metadata` only ever carries generator-produced
        values (dates, counts) — never user free text — but if a free-text
        field (e.g. a custom report title) is ever added here, it MUST go
        through escape_excel_formula() first.
        """
        ws = workbook.create_sheet("Summary Statistics", 0)  # Make it the first sheet

        # Title
        title_font = Font(bold=True, size=14, color='1F4E78')
        ws['A1'] = 'NDAS Research Export - Summary Statistics'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Generation info
        ws['A3'] = 'Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Date Range:'
        ws['B4'] = f"{metadata.get('start_date', 'All')} to {metadata.get('end_date', 'All')}"

        # Record counts
        ws['A7'] = 'Dataset'
        ws['B7'] = 'Record Count'
        ws['C7'] = 'Percentage'
        ws['D7'] = 'Status'
        self.style_header_row(ws, row_num=7)

        row = 8
        total_records = metadata.get('total_records', 0)

        for sheet_name, count in metadata.get('sheets', {}).items():
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count/total_records*100):.1f}%" if total_records > 0 else '0%'
            ws[f'D{row}'] = 'Complete' if count > 0 else 'No Data'

            # Color code status
            if count > 0:
                ws[f'D{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                ws[f'D{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            row += 1

        # Total row
        ws[f'A{row}'] = 'TOTAL'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_records
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = '100%'
        ws[f'C{row}'].font = Font(bold=True)

        # Data Quality Metrics
        row += 3
        ws[f'A{row}'] = 'Data Quality Assessment:'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        ws[f'A{row}'] = 'Quality Metric'
        ws[f'B{row}'] = 'Value'
        ws[f'C{row}'] = 'Status'
        ws[f'D{row}'] = 'Notes'
        self.style_header_row(ws, row_num=row)
        row += 1

        # Quality metrics from metadata
        quality_metrics = metadata.get('quality_metrics', {})

        for metric_name, metric_data in quality_metrics.items():
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = metric_data.get('value', 'N/A')
            ws[f'C{row}'] = metric_data.get('status', 'Unknown')
            ws[f'D{row}'] = metric_data.get('notes', '')

            # Color code quality status
            status = metric_data.get('status', '').lower()
            if status == 'good':
                ws[f'C{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif status == 'warning':
                ws[f'C{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif status == 'critical':
                ws[f'C{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            row += 1

        # General data quality notes
        row += 2
        ws[f'A{row}'] = 'Data Quality Notes:'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = '• All timestamps are in local system time'
        row += 1
        ws[f'A{row}'] = '• Empty cells indicate missing/not applicable data'
        row += 1
        ws[f'A{row}'] = '• For research use, verify data completeness before analysis'
        row += 1
        ws[f'A{row}'] = '• Patient identifiers (BHT, NNC, etc.) may be anonymized per protocol'
        row += 1
        ws[f'A{row}'] = '• Data quality metrics assess completeness and consistency'

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

    def generate(self, output_path=None, start_date=None, end_date=None, parameters=None, institution=None):
        """
        Generate Excel report with selected data

        Args:
            output_path: Optional output file path
            start_date: Optional start date filter
            end_date: Optional end date filter
            parameters: Dict of what to include {'patients': bool, 'gm_assessments': bool, ...}

        Returns:
            tuple: (file_path, metadata dict with counts)
        """
        if parameters is None:
            parameters = {
                'patients': True,
                'gm_assessments': True,
                'hine_assessments': True,
                'developmental_assessments': True,
                'cdic_records': True,
                'gpa_assessments': True,
            }

        # Generate output path
        if output_path is None:
            filename = f"report_{uuid.uuid4()}.xlsx"
            temp_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            output_path = os.path.join(temp_dir, filename)

        # Create workbook
        self.create_workbook()

        # Convert date objects to timezone-aware datetime for filtering
        from datetime import datetime, time
        start_datetime = None
        end_datetime = None
        if start_date:
            start_datetime = timezone.make_aware(datetime.combine(start_date, time.min))
        if end_date:
            end_datetime = timezone.make_aware(datetime.combine(end_date, time.max))

        # Calculate data quality metrics
        quality_metrics = self.calculate_quality_metrics(start_date, end_date, institution=institution)

        metadata = {
            'sheets': {},
            'total_records': 0,
            'start_date': start_date.strftime('%Y-%m-%d') if start_date else 'All',
            'end_date': end_date.strftime('%Y-%m-%d') if end_date else 'All',
            'quality_metrics': quality_metrics
        }

        # Get advanced filters
        filters = parameters.get('filters', {})

        # Add sheets based on parameters (optimized queries)
        if parameters.get('patients', False):
            queryset = Patient.objects.for_institution(institution).select_related('added_by', 'last_edit_by')
            if start_datetime:
                queryset = queryset.filter(created_at__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(created_at__lte=end_datetime)

            # Apply advanced filters
            queryset = self.apply_advanced_filters(queryset, filters, 'patient')

            # Get field selection and anonymization settings
            selected_fields = parameters.get('patient_fields', None)
            anonymize = parameters.get('anonymize_data', False)

            self.add_patients_sheet(self.workbook, queryset, selected_fields, anonymize)
            count = queryset.count()
            metadata['sheets']['Patients'] = count
            metadata['total_records'] += count

        if parameters.get('gm_assessments', False):
            queryset = GMAssessment.objects.select_related('patient', 'added_by').all()
            if start_datetime:
                queryset = queryset.filter(date_of_assessment__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(date_of_assessment__lte=end_datetime)

            if institution:
                queryset = queryset.filter(patient__institution=institution)

            # Apply advanced filters
            queryset = self.apply_advanced_filters(queryset, filters, 'gm')

            self.add_gm_assessments_sheet(self.workbook, queryset)
            count = queryset.count()
            metadata['sheets']['GM Assessments'] = count
            metadata['total_records'] += count

        if parameters.get('hine_assessments', False):
            queryset = HINEAssessment.objects.select_related('patient', 'added_by').all()
            if start_datetime:
                queryset = queryset.filter(date_of_assessment__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(date_of_assessment__lte=end_datetime)

            if institution:
                queryset = queryset.filter(patient__institution=institution)

            # Apply advanced filters
            queryset = self.apply_advanced_filters(queryset, filters, 'hine')

            self.add_hine_assessments_sheet(self.workbook, queryset)
            count = queryset.count()
            metadata['sheets']['HINE Assessments'] = count
            metadata['total_records'] += count

        if parameters.get('developmental_assessments', False):
            queryset = DevelopmentalAssessment.objects.select_related('patient', 'added_by').all()
            if start_datetime:
                queryset = queryset.filter(date_of_assessment__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(date_of_assessment__lte=end_datetime)

            if institution:
                queryset = queryset.filter(patient__institution=institution)

            # Apply advanced filters
            queryset = self.apply_advanced_filters(queryset, filters, 'developmental')

            self.add_developmental_assessments_sheet(self.workbook, queryset)
            count = queryset.count()
            metadata['sheets']['Developmental Assessments'] = count
            metadata['total_records'] += count

        if parameters.get('cdic_records', False):
            queryset = CDICRecord.objects.select_related('patient', 'added_by').all()
            if start_datetime:
                queryset = queryset.filter(assessment_date__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(assessment_date__lte=end_datetime)

            if institution:
                queryset = queryset.filter(patient__institution=institution)

            # Apply advanced filters
            queryset = self.apply_advanced_filters(queryset, filters, 'cdic')

            self.add_cdic_records_sheet(self.workbook, queryset)
            count = queryset.count()
            metadata['sheets']['CDIC Records'] = count
            metadata['total_records'] += count

        if parameters.get('gpa_assessments', False):
            queryset = GeneralPaediatricAssessment.objects.select_related('patient', 'added_by').all()
            if start_datetime:
                queryset = queryset.filter(assessment_date__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(assessment_date__lte=end_datetime)

            if institution:
                queryset = queryset.filter(patient__institution=institution)

            # Apply advanced filters
            queryset = self.apply_advanced_filters(queryset, filters, 'gpa')

            self.add_gpa_assessments_sheet(self.workbook, queryset)
            count = queryset.count()
            metadata['sheets']['GPA Assessments'] = count
            metadata['total_records'] += count

        # Add summary statistics sheet (first sheet)
        self.add_summary_statistics_sheet(self.workbook, metadata)

        # Save workbook
        self.workbook.save(output_path)

        return output_path, metadata

    # ─── Story 2.5 — Institution-scoped aggregate exports ─────────────────────

    def per_institution_aggregate(self, institution):
        """
        Generate a multi-sheet Excel workbook for one institution's full data.
        Sheet 1: Summary | Sheet 2: Patients | Sheet 3+: Assessment types.

        Args:
            institution: Institution model instance (the target institution)

        Returns:
            (output_path, metadata) tuple
        """
        from institution.models import Institution as InstitutionModel

        self.workbook = Workbook()
        # Remove default empty sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

        # ── Patients sheet ─────────────────────────────────────────────────
        patients_qs = Patient.objects.filter(
            institution=institution
        ).select_related('added_by')

        ws_patients = self.workbook.create_sheet("Patients")
        patient_headers = ['ID', 'Name', 'BHT', 'NNC', 'Gender', 'DOB', 'Gestational Age (wks)', 'Created At']
        for col_num, header in enumerate(patient_headers, 1):
            cell = ws_patients.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        for row_num, patient in enumerate(patients_qs, 2):
            ws_patients.cell(row=row_num, column=1, value=patient.id)
            ws_patients.cell(row=row_num, column=2, value=patient.baby_name)
            ws_patients.cell(row=row_num, column=3, value=patient.bht)
            ws_patients.cell(row=row_num, column=4, value=patient.nnc_no)
            ws_patients.cell(row=row_num, column=5, value=patient.sex)
            ws_patients.cell(row=row_num, column=6, value=str(patient.dob_tob) if patient.dob_tob else None)
            ws_patients.cell(row=row_num, column=7, value=patient.pog_wks)
            ws_patients.cell(row=row_num, column=8, value=self.make_naive(patient.created_at))

        # ── Assessment sheets (scoped to this institution's patients) ───────
        patient_ids = patients_qs.values_list('id', flat=True)

        for model_cls, sheet_name in [
            (GMAssessment, 'GMA'),
            (HINEAssessment, 'HINE'),
            (CDICRecord, 'CDIC'),
            (GeneralPaediatricAssessment, 'GPA'),
            (DevelopmentalAssessment, 'DA'),
        ]:
            ws = self.workbook.create_sheet(sheet_name)
            qs = model_cls.objects.filter(patient_id__in=patient_ids).select_related('patient')
            ws.cell(row=1, column=1, value='Patient ID').font = Font(bold=True)
            ws.cell(row=1, column=2, value='Patient Name').font = Font(bold=True)
            ws.cell(row=1, column=3, value='Created At').font = Font(bold=True)
            for row_num, rec in enumerate(qs, 2):
                ws.cell(row=row_num, column=1, value=rec.patient_id)
                ws.cell(row=row_num, column=2, value=rec.patient.baby_name)
                ws.cell(row=row_num, column=3, value=self.make_naive(rec.created_at))

        # ── Summary sheet (first) ─────────────────────────────────────────
        ws_summary = self.workbook.create_sheet("Summary", 0)
        ws_summary.cell(row=1, column=1, value="Per-Institution Report").font = Font(bold=True, size=14)
        ws_summary.cell(row=2, column=1, value=f"Institution: {institution.name}")
        ws_summary.cell(row=3, column=1, value=f"Slug: {institution.slug}")
        ws_summary.cell(row=4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_summary.cell(row=6, column=1, value="Total Patients").font = Font(bold=True)
        ws_summary.cell(row=6, column=2, value=patients_qs.count())

        # ── Save ─────────────────────────────────────────────────────────
        output_dir = os.path.join(settings.MEDIA_ROOT, 'exports', 'institution')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{institution.slug}_{uuid.uuid4().hex}.xlsx")
        self.workbook.save(output_path)

        metadata = {
            'institution': institution.name,
            'total_records': patients_qs.count(),
        }
        return output_path, metadata

    def cross_institution_aggregate(self):
        """
        Generate an aggregate summary workbook across ALL institutions.
        One row per institution with patient count, assessment volumes, subscription status.

        Returns:
            (output_path, metadata) tuple
        """
        from institution.models import Institution as InstitutionModel
        from django.db.models import Count

        self.workbook = Workbook()
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

        ws = self.workbook.create_sheet("Network Summary")
        headers = [
            'Institution', 'Slug', 'Subscription Status', 'Is Active',
            'Total Patients', 'Total Users',
            'GMA Assessments', 'HINE Assessments', 'CDIC Records',
            'GPA Assessments', 'DA Assessments',
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Institution-level counts (related_name='users' and 'patients')
        institutions = InstitutionModel.objects.annotate(
            patient_count=Count('patients', distinct=True),
            user_count=Count('users', distinct=True),
        ).order_by('name')

        for row_num, inst in enumerate(institutions, 2):
            patient_ids = Patient.objects.filter(institution=inst).values_list('id', flat=True)
            gma  = GMAssessment.objects.filter(patient_id__in=patient_ids).count()
            hine = HINEAssessment.objects.filter(patient_id__in=patient_ids).count()
            cdic = CDICRecord.objects.filter(patient_id__in=patient_ids).count()
            gpa  = GeneralPaediatricAssessment.objects.filter(patient_id__in=patient_ids).count()
            da   = DevelopmentalAssessment.objects.filter(patient_id__in=patient_ids).count()

            ws.cell(row=row_num, column=1, value=inst.name)
            ws.cell(row=row_num, column=2, value=inst.slug)
            ws.cell(row=row_num, column=3, value=inst.subscription_status)
            ws.cell(row=row_num, column=4, value=inst.is_active)
            ws.cell(row=row_num, column=5, value=inst.patient_count)
            ws.cell(row=row_num, column=6, value=inst.user_count)
            ws.cell(row=row_num, column=7, value=gma)
            ws.cell(row=row_num, column=8, value=hine)
            ws.cell(row=row_num, column=9, value=cdic)
            ws.cell(row=row_num, column=10, value=gpa)
            ws.cell(row=row_num, column=11, value=da)

        # Grand totals row (bold)
        last_data_row = ws.max_row
        totals_row = last_data_row + 1
        ws.cell(row=totals_row, column=1, value='NETWORK TOTAL').font = Font(bold=True)
        ws.cell(row=totals_row, column=2, value='').font = Font(bold=True)
        ws.cell(row=totals_row, column=3, value='').font = Font(bold=True)
        # Sum columns 4-11 from data rows (row 2 to last_data_row)
        for col_idx in range(4, 12):
            col_letter = ws.cell(row=2, column=col_idx).column_letter
            ws.cell(row=totals_row, column=col_idx,
                    value=f'=SUM({col_letter}2:{col_letter}{last_data_row})').font = Font(bold=True)

        ws.cell(row=totals_row + 2, column=1, value="Report Generated").font = Font(bold=True)
        ws.cell(row=totals_row + 2, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # ── AC #2: Per-institution breakdown sheets (Sheets 2..N) ────────────
        for inst in institutions:
            # Excel sheet names max 31 chars
            sheet_name = inst.name[:31]
            ws_inst = self.workbook.create_sheet(sheet_name)

            ws_inst.cell(row=1, column=1, value=f'{inst.name} — Summary').font = Font(bold=True, size=12)
            ws_inst.cell(row=2, column=1, value=f'Subscription: {inst.get_subscription_status_display()}')
            ws_inst.cell(row=3, column=1, value=f'Slug: {inst.slug}')

            # Assessment summary table
            summary_headers = ['Category', 'Count']
            ws_inst.cell(row=5, column=1, value=summary_headers[0]).font = Font(bold=True)
            ws_inst.cell(row=5, column=2, value=summary_headers[1]).font = Font(bold=True)

            patient_count = Patient.objects.filter(institution=inst).count()
            gma_c  = GMAssessment.objects.filter(patient__institution=inst).count()
            hine_c = HINEAssessment.objects.filter(patient__institution=inst).count()
            cdic_c = CDICRecord.objects.filter(patient__institution=inst).count()
            gpa_c  = GeneralPaediatricAssessment.objects.filter(patient__institution=inst).count()
            da_c   = DevelopmentalAssessment.objects.filter(patient__institution=inst).count()

            rows = [
                ('Patients', patient_count),
                ('GMA Assessments', gma_c),
                ('HINE Assessments', hine_c),
                ('CDIC Records', cdic_c),
                ('GPA Assessments', gpa_c),
                ('Developmental Assessments', da_c),
                ('Total Assessments', gma_c + hine_c + cdic_c + gpa_c + da_c),
            ]
            for r_idx, (label, count) in enumerate(rows, 6):
                ws_inst.cell(row=r_idx, column=1, value=label)
                ws_inst.cell(row=r_idx, column=2, value=count)

            ws_inst.column_dimensions['A'].width = 30
            ws_inst.column_dimensions['B'].width = 12

        output_dir = os.path.join(settings.MEDIA_ROOT, 'exports', 'network')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"network_{uuid.uuid4().hex}.xlsx")
        self.workbook.save(output_path)

        metadata = {
            'scope': 'cross_institution',
            'total_institutions': institutions.count(),
        }
        return output_path, metadata
